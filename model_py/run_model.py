from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from crop_model import (
    Parameters,
    SCENARIOS,
    Scenario,
    scenario_rows,
    validation_rows,
    extreme_condition_test1_rows,
    extreme_condition_test2_rows,
    extreme_condition_test3_rows,
    sensitivity_analysis_rows,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_SECTION_FONT = Font(bold=True, color="1F4E79")


def _write_rows_csv(path: Path, rows: Iterable[Dict[str, float]]) -> None:
    rows = list(rows)
    if not rows:
        raise ValueError("no rows to write")
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _append_table(ws, rows: List[Dict], header_style: bool = True) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    if header_style:
        for cell in ws[ws.max_row]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row[h] for h in headers])


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 28)


def write_workbook(path: Path, params: Parameters) -> None:
    wb = Workbook()

    # Validation and scenarios use the caller-supplied params (may override ethanol_yield).
    # Extreme condition tests and sensitivity analysis are always run with the calibrated
    # default parameters (ethanol_yield=0.046) to reproduce the paper's tables exactly.
    calibrated = Parameters()

    # --- Validation sheet ---
    ws = wb.active
    ws.title = "Validation"
    _append_table(ws, validation_rows(params))

    # --- Scenario sheets ---
    for scenario in SCENARIOS:
        ws = wb.create_sheet(scenario.name)
        _append_table(ws, scenario_rows(scenario, params))

    # --- Extreme Condition Test sheets ---
    _write_ect_sheet(wb, "ECT1 - Zero Valorization", extreme_condition_test1_rows(calibrated))
    _write_ect_sheet(wb, "ECT2 - Zero Pop Growth", extreme_condition_test2_rows(calibrated))
    _write_ect_sheet(wb, "ECT3 - Zero Ethanol Yield", extreme_condition_test3_rows(calibrated))

    # --- Sensitivity Analysis sheets (one per scenario) ---
    for scenario in SCENARIOS:
        _write_sensitivity_sheet(wb, f"SA - {scenario.name}", scenario.name, calibrated)

    # Final formatting pass
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        _autofit(ws)

    wb.save(path)


def _write_ect_sheet(wb: Workbook, sheet_title: str, rows: List[Dict]) -> None:
    ws = wb.create_sheet(sheet_title)
    if not rows:
        return

    # Title row
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row[h] for h in headers])

    # Highlight the warm-up period boundary (Year 2008 = post-warmup start)
    for r_idx, row in enumerate(rows, start=2):
        year_val = ws.cell(row=r_idx, column=1).value
        if year_val == 2008:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = _SECTION_FILL
                ws.cell(row=r_idx, column=c_idx).font = _SECTION_FONT


def _write_sensitivity_sheet(wb: Workbook, sheet_title: str, scenario_name: str, params: Parameters) -> None:
    ws = wb.create_sheet(sheet_title)
    rows = sensitivity_analysis_rows(scenario_name, params)
    if not rows:
        return

    headers = list(rows[0].keys())

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows — add section separator styling when parameter group changes
    prev_param = None
    for row in rows:
        ws.append([row[h] for h in headers])
        current_param = row["Parameter"]
        if current_param != prev_param:
            # Light fill for first row of each parameter group
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c_idx).fill = _SECTION_FILL
            prev_param = current_param


def scenario_by_name(name: str) -> Scenario:
    normalized = name.lower().replace("_", " ").replace("-", " ").strip()
    for scenario in SCENARIOS:
        if scenario.name.lower() == normalized:
            return scenario
    valid = ", ".join(s.name for s in SCENARIOS)
    raise argparse.ArgumentTypeError(f"unknown scenario {name!r}; valid values: {valid}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the sugarcane/ethanol/vinasse system dynamics model.")
    parser.add_argument("--output", default="outputs/scenarios_pysd_python.xlsx", help="Output .xlsx or .csv path.")
    parser.add_argument(
        "--sheet",
        choices=["all", "validation", "scenario"],
        default="all",
        help="Write all workbook sheets, only validation, or one scenario CSV.",
    )
    parser.add_argument("--scenario", type=scenario_by_name, default=SCENARIOS[0], help="Scenario for --sheet scenario.")
    parser.add_argument(
        "--ethanol-yield",
        type=float,
        default=0.046,
        help="Ethanol yield. Use 0.046 to reproduce Scenarios.xlsx/Validation; Crop_v3.alp currently stores 0.05.",
    )
    parser.add_argument(
        "--alp-profile",
        action="store_true",
        help="Shortcut for the literal Crop_v3.alp ethanol yield value of 0.05.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ethanol_yield = 0.05 if args.alp_profile else args.ethanol_yield
    params = Parameters(ethanol_yield=ethanol_yield)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)

    if args.sheet == "all":
        if output.suffix.lower() != ".xlsx":
            raise SystemExit("--sheet all requires an .xlsx output")
        write_workbook(output, params)
    elif args.sheet == "validation":
        _write_rows_csv(output, validation_rows(params))
    else:
        _write_rows_csv(output, scenario_rows(args.scenario, params))

    print(output)


if __name__ == "__main__":
    main()
